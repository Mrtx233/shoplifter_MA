import time
import re
import os
import random
import requests
import lxml.html
import pandas as pd
from colorama import Fore, Style
from openpyxl import load_workbook
import logging
from selenium import webdriver
from selenium.webdriver.common.by import By

logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
}

def write_to_excel(file_path, dataframe, new_sheet):
    if not os.path.exists(file_path):
        dataframe.to_excel(file_path, sheet_name=new_sheet, index=False)
    else:
        book = load_workbook(file_path)
        if new_sheet in book.sheetnames:
            existing_data = pd.read_excel(file_path, sheet_name=new_sheet)
            updated_data = pd.concat([existing_data, dataframe], ignore_index=True)
        else:
            updated_data = dataframe
        with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            updated_data.to_excel(writer, sheet_name=new_sheet, index=False)

def extract_links(url, xpath):
    driver = webdriver.Chrome()
    try:
        driver.get(url)
        time.sleep(10)
        elements = driver.find_elements(By.XPATH, xpath)
        links = [element.get_attribute('href') for element in elements]
        return links
    finally:
        driver.quit()

url = 'http://www.81.cn/rw_208598/index.html'
target_xpath = '//*[@id="main-news-list"]/li/a'
returned_final_links = extract_links(url, target_xpath)

url_name = '中国军网'
notification_path = '人物'
base_path = f'E:\\WorkingWord\\公文爬取(7.15-7.19)\\{url_name}'
excel_file_name = f'{url_name}.xlsx'

tit_xpath = '/html/body/div[3]/div/div[2]/div/h1/text()'
alt_tit_xpath = '/html/body/div[3]/h2/text()'  # 备用XPath
menu_xpath = '//ol[@class="breadcrumb hidden-print"]//a//text()'
fb_time_xpath = '/html/body/div[3]/p/span[4]/text()'
ly_name_xpath = '/html/body/div[3]/p/span[1]/text()'

folder_path = os.path.join(base_path, notification_path)
if not os.path.exists(folder_path):
    os.makedirs(folder_path)
    logger.info(f'Created directory: {folder_path}')
excel_file_path = os.path.join(base_path, excel_file_name)

data = pd.DataFrame()

for page in returned_final_links:
    try:
        time.sleep(random.uniform(0.2, 1))
        page_url = page
        page_response = requests.get(url=page_url, headers=headers)
        page_response.encoding = 'utf-8'
        page_tree = lxml.html.fromstring(page_response.text)

        # 提取标题，如果初始tit_xpath为空则使用备用alt_tit_xpath
        tit = page_tree.xpath(tit_xpath)
        if not tit:
            tit = page_tree.xpath(alt_tit_xpath)
        tit = re.sub(r'[\r\n\s\t]*', '', tit[0]) if tit else '无标题'

        menu = '>'.join(page_tree.xpath(menu_xpath))
        fb_time = page_tree.xpath(fb_time_xpath)
        ly_name = page_tree.xpath(ly_name_xpath)
        file_path = f'{folder_path}\\{tit}.docx'

        with open(file_path, 'a+', encoding='utf-8') as file:
            file.write(tit + '\n')
            for p_text in page_tree.xpath('//*[@id="main-news-list"]//p'):
                if p_text.xpath('.//text()') != '':
                    content = ''.join(p_text.xpath('.//text()'))
                    file.write(content + '\n')
                else:
                    content = ''.join(p_text.xpath('./span/text()'))
                    file.write(content + '\n')

        file_size = os.stat(file_path).st_size // 1024
        if file_size == 0:
            os.remove(file_path)
            logger.info(f'{Fore.RED}{tit}.docx {file_size}KB{Style.RESET_ALL}')
        else:
            logger.info(f'{Fore.BLUE}{tit}.docx {file_size}KB{Style.RESET_ALL}')
            dataMap = {'标题': tit, '目录': menu, '发布时间': fb_time, '来源': ly_name, '地址': page_url}
            data = pd.concat([data, pd.DataFrame([dataMap])], ignore_index=True)

    except Exception as e:
        logger.error(f'Error processing page {page_url}: {e}')
        continue

write_to_excel(file_path=excel_file_path, dataframe=data, new_sheet=notification_path)

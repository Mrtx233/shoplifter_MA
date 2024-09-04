import time  # 导入时间模块，用于延时
import re  # 导入正则表达式模块，用于字符串处理
import os  # 导入操作系统模块，用于文件和目录操作
import random  # 导入随机数模块，用于生成随机延时
import requests  # 导入请求模块，用于发送HTTP请求
import lxml.html  # 导入lxml库中的html模块，用于解析HTML文档
import pandas as pd  # 导入Pandas库，用于数据处理
from colorama import Fore, Style  # 导入Colorama模块，用于终端文本着色
from openpyxl import load_workbook  # 导入Openpyxl模块，用于操作Excel文件
import logging  # 导入日志记录模块
from selenium import webdriver  # 导入Selenium模块中的webdriver，用于自动化浏览器操作
from selenium.webdriver.common.by import By  # 导入Selenium模块中的By，用于元素定位

# 配置日志记录格式
logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)  # 获取日志记录器

# 设置请求头，模拟浏览器访问
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
}

# 定义写入Excel文件的函数
def write_to_excel(file_path, dataframe, new_sheet):
    # 如果文件不存在，创建新的Excel文件并写入数据
    if not os.path.exists(file_path):
        dataframe.to_excel(file_path, sheet_name=new_sheet, index=False)
    else:
        # 加载现有的Excel文件
        book = load_workbook(file_path)
        # 检查工作表是否存在
        if new_sheet in book.sheetnames:
            # 如果存在，读取现有数据并追加新数据
            existing_data = pd.read_excel(file_path, sheet_name=new_sheet)
            updated_data = pd.concat([existing_data, dataframe], ignore_index=True)
        else:
            # 如果不存在，直接写入新数据
            updated_data = dataframe
        # 以追加模式写入数据到Excel文件
        with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            updated_data.to_excel(writer, sheet_name=new_sheet, index=False)

# 定义提取链接的函数
def extract_links(url, xpath):
    driver = webdriver.Chrome()  # 启动Chrome浏览器
    try:
        driver.get(url)  # 打开指定的URL
        time.sleep(10)  # 等待页面加载完成
        elements = driver.find_elements(By.XPATH, xpath)  # 使用XPath定位元素
        links = [element.get_attribute('href') for element in elements]  # 提取元素的href属性，即链接
        return links  # 返回提取的链接列表
    finally:
        driver.quit()  # 关闭浏览器

# 指定要爬取的网页URL和目标XPath
url = 'http://www.81.cn/rw_208598/index.html'
target_xpath = '//*[@id="main-news-list"]/li/a'
# 调用extract_links函数提取链接
returned_final_links = extract_links(url, target_xpath)

# 定义相关变量
url_name = '中国军网'
notification_path = '人物'
base_path = f'E:\\WorkingWord\\公文爬取(7.15-7.19)\\{url_name}'
excel_file_name = f'{url_name}.xlsx'

# 定义各种XPath路径
tit_xpath = '/html/body/div[3]/div/div[2]/div/h1/text()'
alt_tit_xpath = '/html/body/div[3]/h2/text()'  # 备用XPath
menu_xpath = '//ol[@class="breadcrumb hidden-print"]//a//text()'
fb_time_xpath = '/html/body/div[3]/p/span[4]/text()'
ly_name_xpath = '/html/body/div[3]/p/span[1]/text()'

# 创建存储文件的目录，如果不存在则创建
folder_path = os.path.join(base_path, notification_path)
if not os.path.exists(folder_path):
    os.makedirs(folder_path)
    logger.info(f'Created directory: {folder_path}')
# 定义Excel文件路径
excel_file_path = os.path.join(base_path, excel_file_name)

# 创建一个空的DataFrame用于存储数据
data = pd.DataFrame()

# 遍历提取的链接
for page in returned_final_links:
    try:
        time.sleep(random.uniform(0.2, 1))  # 随机延时，避免反爬
        page_url = page  # 当前页面的URL
        page_response = requests.get(url=page_url, headers=headers)  # 发送GET请求获取页面内容
        page_response.encoding = 'utf-8'  # 设置响应编码
        page_tree = lxml.html.fromstring(page_response.text)  # 解析HTML文档

        # 提取标题，如果初始tit_xpath为空则使用备用alt_tit_xpath
        tit = page_tree.xpath(tit_xpath)
        if not tit:
            tit = page_tree.xpath(alt_tit_xpath)
        tit = re.sub(r'[\r\n\s\t]*', '', tit[0]) if tit else '无标题'  # 去除标题中的空白字符

        # 提取目录、发布时间和来源
        menu = '>'.join(page_tree.xpath(menu_xpath))
        fb_time = page_tree.xpath(fb_time_xpath)
        ly_name = page_tree.xpath(ly_name_xpath)
        # 定义存储文件的路径
        file_path = f'{folder_path}\\{tit}.docx'

        # 打开文件并写入内容
        with open(file_path, 'a+', encoding='utf-8') as file:
            file.write(tit + '\n')
            for p_text in page_tree.xpath('//*[@id="main-news-list"]//p'):
                if p_text.xpath('.//text()') != '':
                    content = ''.join(p_text.xpath('.//text()'))
                    file.write(content + '\n')
                else:
                    content = ''.join(p_text.xpath('./span/text()'))
                    file.write(content + '\n')

        # 检查文件大小，如果为0KB则删除文件
        file_size = os.stat(file_path).st_size // 1024
        if file_size == 0:
            os.remove(file_path)
            logger.info(f'{Fore.RED}{tit}.docx {file_size}KB{Style.RESET_ALL}')
        else:
            logger.info(f'{Fore.BLUE}{tit}.docx {file_size}KB{Style.RESET_ALL}')
            # 创建数据映射并追加到DataFrame
            dataMap = {'标题': tit, '目录': menu, '发布时间': fb_time, '来源': ly_name, '地址': page_url}
            data = pd.concat([data, pd.DataFrame([dataMap])], ignore_index=True)

    except Exception as e:
        logger.error(f'Error processing page {page_url}: {e}')  # 捕获异常并记录错误
        continue

# 调用write_to_excel函数将数据写入Excel文件
write_to_excel(file_path=excel_file_path, dataframe=data, new_sheet=notification_path)

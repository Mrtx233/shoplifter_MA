import random  # 导入随机数模块，用于生成随机延时
import re  # 导入正则表达式模块，用于处理字符串
import os  # 导入操作系统模块，用于文件和目录操作
import time  # 导入时间模块，用于添加延时
import lxml.html  # 导入lxml库中的html模块，用于解析HTML文档
import pandas as pd  # 导入Pandas库，用于数据处理
import requests  # 导入requests模块，用于发送HTTP请求
from colorama import Fore, Style  # 导入Colorama模块，用于终端文本着色
from bs4 import BeautifulSoup  # 导入BeautifulSoup模块，用于解析HTML
from docx import Document  # 导入python-docx模块，用于操作Word文档
import logging  # 导入日志记录模块，用于记录日志
from openpyxl import load_workbook  # 导入openpyxl模块，用于操作Excel文件

# 配置日志记录器，设置日志级别和格式
logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)  # 获取日志记录器实例

# 设置请求头，模拟浏览器访问
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
}

# 定义将数据写入Excel文件的函数
def write_to_excel(file_path, dataframe, new_sheet):
    if not os.path.exists(file_path):  # 如果文件不存在
        dataframe.to_excel(file_path, sheet_name=new_sheet, index=False)  # 创建新的Excel文件并写入数据
    else:
        book = load_workbook(file_path)  # 加载现有的Excel文件
        if new_sheet in book.sheetnames:  # 如果工作表存在
            existing_data = pd.read_excel(file_path, sheet_name=new_sheet)  # 读取现有的数据
            updated_data = pd.concat([existing_data, dataframe], ignore_index=True)  # 合并现有数据和新数据
        else:
            updated_data = dataframe  # 如果工作表不存在，直接使用新数据
        with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:  # 以追加模式打开Excel文件
            updated_data.to_excel(writer, sheet_name=new_sheet, index=False)  # 将数据写入工作表

# 设置基础URL和文件路径相关变量
base_url = 'http://sfj.wuzhou.gov.cn/jcdt/'
url_name = '梧州市司法局'
notification_path = '基层动态'
base_path = f'E:\\WorkingWord\\公文爬取(7.15-7.19)\\司法局\\广西壮族自治区\\{url_name}'
excel_file_name = f'{url_name}.xlsx'

# 定义XPath路径
page_list_xpath = '//*[@id="morelist"]/ul/li/a/@href'
tit_xpath = '/html/body/div[11]/div[2]/div/h1/text()'
menu_xpath = '//div[@class="crumb-nav"]//a//text()'
fb_time_xpath = '/html/body/div[11]/div[2]/div/div[1]/div[1]/text()'
ly_name_xpath = '/html/body/div[11]/div[2]/div/div[1]/div[1]/text()'

# 创建存储文件的目录，如果不存在则创建
folder_path = os.path.join(base_path, notification_path)
excel_file_path = os.path.join(base_path, excel_file_name)
if not os.path.exists(folder_path):
    os.makedirs(folder_path)
    logger.info(f'Created directory: {folder_path}')

# 创建一个空的DataFrame用于存储数据
data = pd.DataFrame()

# 循环抓取页面数据
for i in range(0, 2):  # 遍历分页
    if i == 0:
        url = base_url  # 第一页的URL
    else:
        url = f'{base_url}index_{i}.html'  # 后续页面的URL

    logger.info(f'URL: {url}')  # 记录当前抓取的URL
    try:
        response = requests.get(url=url, headers=headers)  # 发送GET请求获取页面内容
        response.raise_for_status()  # 如果响应状态码不是200，则抛出异常
        response.encoding = 'utf-8'  # 设置响应编码
        tree = lxml.html.fromstring(response.text)  # 解析HTML文档
        page_list = tree.xpath(page_list_xpath)  # 使用XPath提取页面列表

        for page in page_list:  # 遍历页面列表
            try:
                time.sleep(random.uniform(1, 1.5))  # 随机延时，避免反爬
                page_url = base_url + re.sub(r'\./', '/', page)  # 生成页面URL
                page_response = requests.get(url=page_url, headers=headers)  # 发送GET请求获取页面内容
                page_response.raise_for_status()  # 如果响应状态码不是200，则抛出异常
                page_response.encoding = 'utf-8'  # 设置响应编码
                page_tree = lxml.html.fromstring(page_response.text)  # 解析HTML文档

                file_name = re.sub(r'[\r\n\s\t\|]*', '', page_tree.xpath(tit_xpath)[0])  # 提取标题并去除无效字符
                menu = '>'.join(page_tree.xpath(menu_xpath))  # 提取目录并拼接成字符串
                fb_time = page_tree.xpath(fb_time_xpath)  # 提取发布时间
                ly_name = page_tree.xpath(ly_name_xpath)  # 提取来源

                # 使用BeautifulSoup解析HTML内容
                html_content = page_response.content
                soup = BeautifulSoup(html_content, 'html.parser')
                target_div = soup.find('div', class_='article-con')  # 查找目标div
                text_content = target_div.get_text()  # 提取文本内容

                # 将内容写入Word文档
                doc = Document()
                doc.add_paragraph(text_content)
                save_path = os.path.join(folder_path, f'{file_name}.docx')  # 生成保存路径
                doc.save(save_path)  # 保存Word文档

                file_size = os.stat(save_path).st_size // 1024  # 获取文件大小（KB）
                if file_size == 0:
                    os.remove(save_path)  # 如果文件大小为0KB，则删除文件
                    logger.info(f'{Fore.RED}{file_name}.docx {file_size}KB{Style.RESET_ALL}')
                else:
                    logger.info(f'{Fore.BLUE}{file_name}.docx {file_size}KB{Style.RESET_ALL}')
                    data_map = {'标题': file_name, '目录': menu, '发布时间': fb_time, '来源': ly_name, '地址': page_url}  # 创建数据映射
                    data = pd.concat([data, pd.DataFrame([data_map])], ignore_index=True)  # 将数据追加到DataFrame

            except Exception as e:
                logger.error(f"Error processing page {page_url}: {e}")  # 捕获并记录处理页面时的异常

    except Exception as e:
        logger.error(f"Error fetching URL {url}: {e}")  # 捕获并记录获取URL时的异常

# 将数据写入Excel文件
write_to_excel(file_path=excel_file_path, dataframe=data, new_sheet=notification_path)

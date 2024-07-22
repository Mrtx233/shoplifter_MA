import random  # 导入随机数模块，用于生成随机延时
import re  # 导入正则表达式模块，用于处理字符串
import os  # 导入操作系统模块，用于文件和目录操作
import time  # 导入时间模块，用于添加延时
import lxml.html  # 导入lxml库中的html模块，用于解析HTML文档
import pandas as pd  # 导入Pandas库，用于数据处理
import requests  # 导入requests模块，用于发送HTTP请求
from colorama import Fore, Style  # 导入Colorama模块，用于终端文本着色
from bs4 import BeautifulSoup  # 导入BeautifulSoup模块，用于解析HTML
from docx import Document  # 导入python-docx模块，用于操作Word文档
import logging  # 导入日志记录模块，用于记录日志
from openpyxl import load_workbook  # 导入openpyxl模块，用于操作Excel文件
from docx.shared import Pt  # 导入Pt模块，用于设置字体大小
from docx.oxml.ns import qn  # 导入qn模块，用于设置字体

# 配置日志记录器，设置日志级别和格式
logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)  # 获取日志记录器实例

# 设置请求头，模拟浏览器访问
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
}

# 定义将数据写入Excel文件的函数
def write_to_excel(file_path, dataframe, new_sheet):
    if not os.path.exists(file_path):  # 如果文件不存在
        dataframe.to_excel(file_path, sheet_name=new_sheet, index=False)  # 创建新的Excel文件并写入数据
    else:
        book = load_workbook(file_path)  # 加载现有的Excel文件
        if new_sheet in book.sheetnames:  # 如果工作表存在
            existing_data = pd.read_excel(file_path, sheet_name=new_sheet)  # 读取现有的数据
            updated_data = pd.concat([existing_data, dataframe], ignore_index=True)  # 合并现有数据和新数据
        else:
            updated_data = dataframe  # 如果工作表不存在，直接使用新数据
        with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:  # 以追加模式打开Excel文件
            updated_data.to_excel(writer, sheet_name=new_sheet, index=False)  # 将数据写入工作表

# 定义递归函数，打印标签及其内容到Word文档
def print_tags(tag, level=0, doc=None):
    children = tag.find_all(recursive=False)  # 查找直接子标签
    if not children:  # 如果没有子标签
        text = tag.get_text(strip=True)  # 提取标签文本
        if text:  # 如果文本不为空
            if doc:  # 如果文档对象存在
                paragraph = doc.add_paragraph(text)  # 将文本添加到Word文档中
                run = paragraph.runs[0]  # 获取文档中的第一个运行（文本段）
                run.font.name = 'Times New Roman'  # 设置字体为Times New Roman
                run._element.rPr.rFonts.set(qn('w:eastAsia'), 'Times New Roman')  # 设置字体为Times New Roman（中文）
                run.font.size = Pt(12)  # 设置字体大小为12磅
    for child in children:  # 遍历子标签
        print_tags(child, level + 1, doc)  # 递归调用，处理子标签

# 主函数，抓取网页内容并将内容保存到Word文档
def main(url, target_tag, target_attr=None, output_path='output.docx'):
    response = requests.get(url)  # 发送GET请求获取页面内容
    response.encoding = 'utf-8'  # 设置响应编码
    if response.status_code != 200:  # 如果响应状态码不是200
        logger.error(f"无法访问网站：{url}")  # 记录错误日志
        return  # 返回，不继续执行
    soup = BeautifulSoup(response.text, 'html.parser')  # 使用BeautifulSoup解析HTML内容
    tags = soup.find_all(target_tag, target_attr) if target_attr else soup.find_all(target_tag)  # 查找目标标签
    doc = Document()  # 创建一个新的Word文档
    for tag in tags:  # 遍历目标标签
        print_tags(tag, doc=doc)  # 将标签内容打印到Word文档中
    os.makedirs(os.path.dirname(output_path), exist_ok=True)  # 创建目录（如果不存在）
    doc.save(output_path)  # 保存Word文档到指定路径
    # logger.info(f"文档已保存到 {output_path}")  # 记录文档保存的路径（注释掉了）

# 设置基础URL和文件路径相关变量
base_url = 'http://www.beihai.gov.cn/xxgkbm/bhsmzj/gzxx_34/'
url_name = '北海市民政局'
notification_path = '工作信息'
base_path = f'E:\\WorkingWord\\公文爬取(7.15-7.19)\\民政局\\广西壮族自治区\\{url_name}'
excel_file_name = f'{url_name}.xlsx'

# 定义XPath路径
page_list_xpath = '//*[@id="morelist"]/ul/li/a/@href'
tit_xpath = '/html/body/div[11]/div[2]/div/h1/text()'
menu_xpath = '//div[@class="crumb-nav"]//a//text()'
fb_time_xpath = '/html/body/div[11]/div[2]/div/div[1]/div[1]/text()'
ly_name_xpath = '/html/body/div[11]/div[2]/div/div[1]/div[1]/text()'

# 定义BeautifulSoup抓取目标的标签和属性
target_tag = 'div'
target_attr = {'class': 'article-con'}

# 创建存储文件的目录，如果不存在则创建
folder_path = os.path.join(base_path, notification_path)
excel_file_path = os.path.join(base_path, excel_file_name)
if not os.path.exists(folder_path):
    os.makedirs(folder_path)
    logger.info(f'Created directory: {folder_path}')

# 创建一个空的DataFrame用于存储数据
data = pd.DataFrame()

# 循环抓取页面数据
for i in range(2, 17):  # 遍历分页
    if i == 0:
        url = base_url  # 第一页的URL
    else:
        url = f'{base_url}index_{i}.shtml'  # 后续页面的URL
    logger.info(f'URL: {url}')  # 记录当前抓取的URL
    try:
        response = requests.get(url=url, headers=headers)  # 发送GET请求获取页面内容
        response.raise_for_status()  # 如果响应状态码不是200，则抛出异常
        response.encoding = 'utf-8'  # 设置响应编码
        tree = lxml.html.fromstring(response.text)  # 解析HTML文档
        page_list = tree.xpath(page_list_xpath)  # 使用XPath提取页面列表
        for page in page_list:  # 遍历页面列表
            try:
                time.sleep(random.uniform(1, 1.5))  # 随机延时，避免反爬
                page_url = base_url + re.sub(r'\./', '/', page)  # 生成页面URL
                page_response = requests.get(url=page_url, headers=headers)  # 发送GET请求获取页面内容
                page_response.raise_for_status()  # 如果响应状态码不是200，则抛出异常
                page_response.encoding = 'utf-8'  # 设置响应编码
                page_tree = lxml.html.fromstring(page_response.text)  # 解析HTML文档
                file_name = re.sub(r'[\r\n\s\t\|]*', '', page_tree.xpath(tit_xpath)[0])  # 提取标题并去除无效字符
                menu = '>'.join(page_tree.xpath(menu_xpath))  # 提取目录并拼接成字符串
                fb_time = page_tree.xpath(fb_time_xpath)  # 提取发布时间
                ly_name = page_tree.xpath(ly_name_xpath)  # 提取来源
                # 调用main函数抓取页面内容并保存为Word文档
                main(page_url, target_tag, target_attr, os.path.join(folder_path, f'{file_name}.docx'))
                # 获取Word文档大小（KB）
                file_size = os.stat(os.path.join(folder_path, f'{file_name}.docx')).st_size // 1024
                if file_size < 35:  # 如果文件大小小于35KB
                    os.remove(os.path.join(folder_path, f'{file_name}.docx'))  # 删除文件
                    logger.info(f'{Fore.RED}{file_name}.docx {file_size}KB{Style.RESET_ALL}')
                else:
                    logger.info(f'{Fore.BLUE}{file_name}.docx {file_size}KB{Style.RESET_ALL}')
                    data_map = {'标题': file_name, '目录': menu, '发布时间': fb_time, '来源': ly_name, '地址': page_url}  # 创建数据映射
                    data = pd.concat([data, pd.DataFrame([data_map])], ignore_index=True)  # 将数据追加到DataFrame
            except Exception as e:
                logger.error(f"Error processing page {page_url}: {e}")  # 捕获并记录处理页面时的异常
    except Exception as e:
        logger.error(f"Error fetching URL {url}: {e}")  # 捕获并记录获取URL时的异常

# 将数据写入Excel文件
write_to_excel(file_path=excel_file_path, dataframe=data, new_sheet=notification_path)

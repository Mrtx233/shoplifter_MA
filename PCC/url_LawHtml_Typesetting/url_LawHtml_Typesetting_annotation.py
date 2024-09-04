import random  # 导入用于生成随机数的模块
import re  # 导入正则表达式模块，用于处理字符串
import os  # 导入操作系统模块，用于文件和目录操作
import time  # 导入时间模块，用于添加延时
import lxml.html  # 导入lxml库中的html模块，用于解析HTML文档
import pandas as pd  # 导入Pandas库，用于数据处理
import requests  # 导入requests模块，用于发送HTTP请求
from colorama import Fore, Style  # 导入Colorama模块，用于终端文本着色
from openpyxl import load_workbook  # 导入openpyxl模块，用于操作Excel文件
import logging  # 导入日志记录模块，用于记录日志

# 设置日志记录系统
logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

# 定义HTTP请求的头部信息
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
}

# 定义将DataFrame数据写入Excel文件的函数
def write_to_excel(file_path, dataframe, new_sheet):
    if not os.path.exists(file_path):
        dataframe.to_excel(file_path, sheet_name=new_sheet, index=False)  # 文件不存在时，直接写入新文件
    else:
        book = load_workbook(file_path)  # 加载现有的Excel文件
        if new_sheet in book.sheetnames:  # 检查指定的工作表是否存在
            existing_data = pd.read_excel(file_path, sheet_name=new_sheet)  # 读取现有工作表的数据
            updated_data = pd.concat([existing_data, dataframe], ignore_index=True)  # 合并现有数据和新数据
        else:
            updated_data = dataframe  # 如果工作表不存在，则直接使用新数据
        with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            updated_data.to_excel(writer, sheet_name=new_sheet, index=False)  # 追加写入数据到指定工作表

# 基础URL和路径设置
base_url = 'https://www.qxn.gov.cn/zwgk/zfjg/zgaj_5135056/bmxxgkml_5135059/hjfw/'
url_name = '甘肃省民政厅'
notification_path = '甘肃要闻'
base_path = f'E:\\WorkingWord\\公文爬取(7.15-7.19)\\民政局\\陕西省\\{url_name}'
excel_file_name = f'{url_name}.xlsx'

# 定义XPath路径
page_list_xpath = '//*[@id="main"]/div[1]/div[2]/div/div[2]/ul/li/a/@href'
tit_xpath = '//*[@id="ConBox"]/div[2]/h1/text()'
menu_xpath = '//div[@class="dh mb5 b-ltrb"]//a//text()'
fb_time_xpath = '//*[@id="ConBox"]/div[2]/div/div[1]/span[2]/text()'
ly_name_xpath = '//*[@id="ConBox"]/div[2]/div/div[1]/span[1]/text()'

# 创建存储文件的目录
folder_path = os.path.join(base_path, notification_path)
if not os.path.exists(folder_path):
    os.makedirs(folder_path)
    logger.info(f'Created directory: {folder_path}')

# 初始化Excel文件路径和数据存储
excel_file_path = os.path.join(base_path, excel_file_name)
data = pd.DataFrame()

# 循环抓取分页数据
for i in range(0, 2):
    if i == 0:
        url = base_url  # 第一页URL
    else:
        url = f'{base_url}index_{i}.html'  # 其他页的URL

    logger.info(f'URL: {url}')
    response = requests.get(url=url, headers=headers)  # 发送HTTP请求获取页面内容
    response.raise_for_status()  # 检查请求是否成功，抛出异常如果失败
    response.encoding = 'utf-8'  # 设置响应编码
    tree = lxml.html.fromstring(response.text)  # 解析HTML文档
    page_list = tree.xpath(page_list_xpath)  # 提取页面列表中的链接

    for page in page_list:
        try:
            time.sleep(random.uniform(1, 1.5))  # 随机延时，防止过于频繁的请求
            page_url = page  # 获取页面链接
            # page_url = base_url + re.sub(r'\./', '/', page)  # 如果需要处理相对路径，可以取消注释
            page_response = requests.get(url=page_url, headers=headers)  # 发送HTTP请求获取每个页面的内容
            page_response.encoding = 'utf-8'  # 设置响应编码
            page_tree = lxml.html.fromstring(page_response.text)  # 解析HTML文档

            tit = re.sub(r'[\r\n\s\t]*', '', page_tree.xpath(tit_xpath)[0])  # 提取标题并去除多余字符
            menu = '>'.join(page_tree.xpath(menu_xpath))  # 提取面包屑导航的文本
            fb_time = page_tree.xpath(fb_time_xpath)  # 提取发布时间
            ly_name = page_tree.xpath(ly_name_xpath)  # 提取来源名称
            file_path = f'{folder_path}\\{tit}.docx'  # 生成文档保存路径

            with open(file_path, 'a+', encoding='utf-8') as file:
                file.write(tit + '\n')  # 写入标题
                for p_text in page_tree.xpath('//div[@id="Zoom"]//p'):  # 提取内容
                    if p_text.xpath('.//text()') != '':
                        content = ''.join(p_text.xpath('.//text()'))
                        file.write(content + '\n')  # 写入段落内容
                    else:
                        content = ''.join(p_text.xpath('./span/text()'))
                        file.write(content + '\n')  # 写入其他内容

            file_size = os.stat(file_path).st_size // 1024  # 获取文件大小（KB）
            if file_size == 0:
                os.remove(file_path)  # 删除空文件
                logger.info(f'{Fore.RED}{tit}.docx {file_size}KB{Style.RESET_ALL}')
            else:
                logger.info(f'{Fore.BLUE}{tit}.docx {file_size}KB{Style.RESET_ALL}')
                dataMap = {'标题': tit, '目录': menu, '发布时间': fb_time, '来源': ly_name, '地址': page_url}
                data = pd.concat([data, pd.DataFrame([dataMap])], ignore_index=True)  # 将数据添加到DataFrame中

        except Exception as e:
            logger.error(f'Error processing page {page_url}: {e}')  # 记录处理页面时的错误
            continue

# 将数据写入Excel文件
write_to_excel(file_path=excel_file_path, dataframe=data, new_sheet=notification_path)

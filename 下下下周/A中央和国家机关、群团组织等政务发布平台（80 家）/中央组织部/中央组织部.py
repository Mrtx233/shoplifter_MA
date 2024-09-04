import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import time
import re
import os
import requests
import lxml.html
import pandas as pd
from colorama import Fore, Style
from openpyxl import load_workbook
import logging
from selenium import webdriver
from selenium.webdriver.common.by import By

logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
}

def write_to_excel(file_path, dataframe, new_sheet):
    if not os.path.exists(file_path):
        dataframe.to_excel(file_path, sheet_name=new_sheet, index=False)
    else:
        book = load_workbook(file_path)
        if new_sheet in book.sheetnames:
            existing_data = pd.read_excel(file_path, sheet_name=new_sheet)
            updated_data = pd.concat([existing_data, dataframe], ignore_index=True)
        else:
            updated_data = dataframe
        with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            updated_data.to_excel(writer, sheet_name=new_sheet, index=False)

def extract_links_from_all_pages(driver, link_xpath):
    elements = driver.find_elements(By.XPATH, link_xpath)
    links = [element.get_attribute('href') for element in elements]
    return links

def scrape_links(url, link_xpath, next_button_xpath, click_times):
    driver = webdriver.Chrome()
    all_links = []

    try:
        driver.get(url)

        for _ in range(click_times):
            time.sleep(1)  # 等待页面加载

            # 点击下一页按钮
            next_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, next_button_xpath))
            )
            next_button.click()

        # 提取所有页面的链接
        time.sleep(10)  # 确保最后一页完全加载
        all_links = extract_links_from_all_pages(driver, link_xpath)

    finally:
        driver.quit()

    return all_links

# 示例用法
url = 'https://news.12371.cn/dzybmbdj/zzb/ldhd/'  # 替换为实际网站
link_xpath = '//*[@id="active_md"]/ul/li/h3/a'  # 替换为实际包含链接的div的XPath
next_button_xpath = '//*[@class="one_btn_previous"]'  # 替换为实际的“下一页”按钮的XPath
click_times = 30  # 替换为需要点击“下一页”的次数

all_links = scrape_links(url, link_xpath, next_button_xpath, click_times)
url_name = '中央组织部'
notification_path = '组织工作-领导活动'
base_path = f'E:\\WorkingWord\\马缕_新闻稿采集(7.22-7.26)\\中央和国家机关、群团组织等政务发布平台\\{url_name}'
excel_file_name = f'{url_name}.xlsx'

tit_xpath = '//*[@id="page_body"]/div[5]/div[4]/div/div[1]/h1/text()'
menu_xpath = '//div[@class="catpos"]//a//text()'
fb_time_xpath = '//*[@id="page_body"]/div[5]/div[4]/div/div[6]/i/text()'
ly_name_xpath = '//*[@id="page_body"]/div[5]/div[4]/div/div[6]/i/text()'

folder_path = os.path.join(base_path, notification_path)
if not os.path.exists(folder_path):
    os.makedirs(folder_path)
    logger.info(f'Created directory: {folder_path}')
excel_file_path = os.path.join(base_path, excel_file_name)

data = pd.DataFrame()

# 输出所有链接
for link in all_links:
    try:
        # time.sleep(random.uniform(0.5, 0.6))
        page_url = link
        page_response = requests.get(url=page_url, headers=headers)
        page_response.encoding = 'utf-8'
        page_tree = lxml.html.fromstring(page_response.text)

        tit = re.sub(r'[\r\n\s\t]*', '', page_tree.xpath(tit_xpath)[0])
        menu = '>'
        fb_time = page_tree.xpath(fb_time_xpath)
        ly_name = page_tree.xpath(ly_name_xpath)
        file_path = f'{folder_path}\\{tit}.docx'

        with open(file_path, 'a+', encoding='utf-8') as file:
            file.write(tit + '\n')
            for p_text in page_tree.xpath('//*[@id="font_area"]/div/div//p'):
                if p_text.xpath('.//text()') != '':
                    content = ''.join(p_text.xpath('.//text()'))
                    file.write(content + '\n')
                else:
                    content = ''.join(p_text.xpath('./span/text()'))
                    file.write(content + '\n')

        file_size = os.stat(file_path).st_size // 1024
        if file_size == 0:
            os.remove(file_path)
            logger.info(f'{Fore.RED}{tit}.docx {file_size}KB{Style.RESET_ALL}')
        else:
            logger.info(f'{Fore.BLUE}{tit}.docx {file_size}KB{Style.RESET_ALL}')
            dataMap = {'标题': tit, '目录': menu, '发布时间': fb_time, '来源': ly_name, '地址': page_url}
            data = pd.concat([data, pd.DataFrame([dataMap])], ignore_index=True)

    except Exception as e:
        logger.error(f'Error processing page {page_url}: {e}')
        continue

write_to_excel(file_path=excel_file_path, dataframe=data, new_sheet=notification_path)


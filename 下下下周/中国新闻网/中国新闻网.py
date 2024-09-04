import random
import re
import os
import time
import lxml.html
import pandas as pd
import requests
from colorama import Fore, Style
from openpyxl import load_workbook
import logging

logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
}

def write_to_excel(file_path, dataframe, new_sheet):
    if not os.path.exists(file_path):
        dataframe.to_excel(file_path, sheet_name=new_sheet, index=False)
    else:
        book = load_workbook(file_path)
        if new_sheet in book.sheetnames:
            existing_data = pd.read_excel(file_path, sheet_name=new_sheet)
            updated_data = pd.concat([existing_data, dataframe], ignore_index=True)
        else:
            updated_data = dataframe
        with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            updated_data.to_excel(writer, sheet_name=new_sheet, index=False)

url_name = '中国新闻网'
# A_url = 'https://it.gmw.cn/'
base_urls = [
    'https://channel.chinanews.com.cn/cns/cl/gn-gcdt.shtml',
    'https://channel.chinanews.com.cn/cns/cl/gn-rsbd.shtml',
    'https://channel.chinanews.com.cn/cns/cl/fz-ffcl.shtml',
    'https://channel.chinanews.com.cn/u/sdbd.shtml',
    'https://channel.chinanews.com.cn/u/gn-la.shtml',
    'https://channel.chinanews.com.cn/cns/cl/gn-kjww.shtml',
    'https://channel.chinanews.com.cn/cns/cl/gn-js.shtml'
]

notification_paths = [
    '时政-高层',
    '时政-人事',
    '时政-反腐',
    '时政-深度',
    '时政-两岸',
    '时政-科教',
    '时政-军事'
]


page_list_xpath = '//*[@id="ent0"]/li/div[1]/a/@href'
tit_xpath = '//*[@id="cont_1_1_2"]/div[2]/h1/text()'
menu_xpath = '//div[@class="g-crumbs"]//a//text()'
fb_time_xpath = '//*[@id="cont_1_1_2"]/div[2]/div[2]/text()'
ly_name_xpath = '//*[@id="cont_1_1_2"]/div[2]/div[2]/a/text()'
base_path = f'E:\\WorkingWord\\马缕_新闻稿采集(7.22-7.26)\\中央新闻网站\\{url_name}'
excel_file_name = f'{url_name}.xlsx'

for base_url, notification_path in zip(base_urls, notification_paths):
    folder_path = os.path.join(base_path, notification_path)
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
        logger.info(f'Created directory: {folder_path}')
    excel_file_path = os.path.join(base_path, excel_file_name)
    data = pd.DataFrame()

    for i in range(1, 11):
        try:
            if i == 1:
                url = f'{base_url}'
            else:
                # https://channel.chinanews.com.cn/cns/cl/gn-gcdt.shtml?pager=1
                url = f'{base_url}?pager=i'
            logger.info(f'URL: {url}')
            response = requests.get(url=url, headers=headers)
            response.raise_for_status()
            response.encoding = 'utf-8'
            tree = lxml.html.fromstring(response.text)
            page_list = tree.xpath(page_list_xpath)

            for page in page_list:
                try:
                    # time.sleep(random.uniform(0.5, 0.8))
                    page_url = page
                    # page_url = A_url + re.sub(r'\./', '/', page)
                    page_response = requests.get(url=page_url, headers=headers)
                    page_response.encoding = 'utf-8'
                    page_tree = lxml.html.fromstring(page_response.text)
                    tit = re.sub(r'[\r\n\s\t\|]*', '', page_tree.xpath(tit_xpath)[0])
                    menu = '>'.join(page_tree.xpath(menu_xpath))
                    fb_time = page_tree.xpath(fb_time_xpath)
                    ly_name = page_tree.xpath(ly_name_xpath)
                    file_path = f'{folder_path}\\{tit}.docx'

                    with open(file_path, 'a+', encoding='utf-8') as file:
                        file.write(tit + '\n')
                        for p_text in page_tree.xpath('//*[@id="cont_1_1_2"]/div[2]/div[4]/div[2]//p'):
                            if p_text.xpath('.//text()') != '':
                                content = ''.join(p_text.xpath('.//text()'))
                                file.write(content + '\n')
                            else:
                                content = ''.join(p_text.xpath('./span/text()'))
                                file.write(content + '\n')

                    file_size = os.stat(file_path).st_size // 1024
                    if file_size == 0:
                        os.remove(file_path)
                        logger.info(f'{Fore.RED}{tit}.docx {file_size}KB{Style.RESET_ALL}')
                    else:
                        logger.info(f'{Fore.BLUE}{tit}.docx {file_size}KB{Style.RESET_ALL}')
                        dataMap = {'标题': tit, '目录': menu, '发布时间': fb_time, '来源': ly_name, '地址': page_url}
                        data = pd.concat([data, pd.DataFrame([dataMap])], ignore_index=True)

                except Exception as e:
                    logger.error(f'Error processing page {page_url}: {e}')
                    continue
        except Exception as e:
            logger.error(f'Error processing page {page_url}: {e}')
            continue

    write_to_excel(file_path=excel_file_path, dataframe=data, new_sheet=notification_path)
    time.sleep(10)
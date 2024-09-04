import random
import re
import os
import time
import lxml.html
import pandas as pd
import requests
from colorama import Fore, Style
from bs4 import BeautifulSoup
from docx import Document
import logging
from openpyxl import load_workbook
from docx.shared import Pt
from docx.oxml.ns import qn

logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
}

def write_to_excel(file_path, dataframe, new_sheet):
    if not os.path.exists(file_path):
        dataframe.to_excel(file_path, sheet_name=new_sheet, index=False)
    else:
        book = load_workbook(file_path)
        if new_sheet in book.sheetnames:
            existing_data = pd.read_excel(file_path, sheet_name=new_sheet)
            updated_data = pd.concat([existing_data, dataframe], ignore_index=True)
        else:
            updated_data = dataframe
        with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            updated_data.to_excel(writer, sheet_name=new_sheet, index=False)

def print_tags(tag, level=0, doc=None):
    children = tag.find_all(recursive=False)
    if not children:
        text = tag.get_text(strip=True)
        if text:
            if doc:
                paragraph = doc.add_paragraph(text)
                run = paragraph.runs[0]
                run.font.name = 'Times New Roman'
                run._element.rPr.rFonts.set(qn('w:eastAsia'), 'Times New Roman')
                run.font.size = Pt(12)
    for child in children:
        print_tags(child, level + 1, doc)

def main(url, target_tag, target_attr=None, output_path='output.docx'):
    response = requests.get(url)
    response.encoding = 'utf-8'
    if response.status_code != 200:
        logger.error(f"无法访问网站：{url}")
        return
    soup = BeautifulSoup(response.text, 'html.parser')
    tags = soup.find_all(target_tag, target_attr) if target_attr else soup.find_all(target_tag)
    doc = Document()
    for tag in tags:
        print_tags(tag, doc=doc)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    doc.save(output_path)
    # logger.info(f"文档已保存到 {output_path}")

base_url = 'https://www.ceweekly.cn/finance/banking/'
url_name = '中国经济周刊网(经济网)'
notification_path = '金融'
base_path = f'E:\\WorkingWord\\马缕_公文爬取(7.15-7.19)\\{url_name}'
excel_file_name = f'{url_name}.xlsx'

page_list_xpath = '/html/body/div[2]/div[1]/div[2]/div/div/div[1]/div[1]/a/@href'
tit_xpath = '/html/body/div[2]/div/div[1]/div/div[3]/text()'
menu_xpath = '//div[@class="breadcrumbs mb-30"]//a//text()'
fb_time_xpath = '/html/body/div[2]/div/div[1]/div/div[5]/div[2]/text()'
ly_name_xpath = '/html/body/div[2]/div/div[1]/div/div[5]/div[1]/text()'
target_tag = 'div'
target_attr = {'class': 'wrap article'}
folder_path = os.path.join(base_path, notification_path)
excel_file_path = os.path.join(base_path, excel_file_name)

if not os.path.exists(folder_path):
    os.makedirs(folder_path)
    logger.info(f'Created directory: {folder_path}')

data = pd.DataFrame()

for i in range(1, 11):
    if i == 1:
        url = base_url
    else:
        url = f'{base_url}/{i}.html'
    logger.info(f'URL: {url}')
    try:
        response = requests.get(url=url, headers=headers)
        response.raise_for_status()
        response.encoding = 'utf-8'
        tree = lxml.html.fromstring(response.text)
        page_list = tree.xpath(page_list_xpath)
        for page in page_list:
            try:
                time.sleep(random.uniform(1, 1.5))
                page_url =page
                # page_url = base_url + re.sub(r'\./', '/', page)
                page_response = requests.get(url=page_url, headers=headers)
                page_response.raise_for_status()
                page_response.encoding = 'utf-8'
                page_tree = lxml.html.fromstring(page_response.text)
                file_name = re.sub(r'[\r\n\s\t\|]*', '', page_tree.xpath(tit_xpath)[0])
                menu = '>'.join(page_tree.xpath(menu_xpath))
                fb_time = page_tree.xpath(fb_time_xpath)
                ly_name = page_tree.xpath(ly_name_xpath)
                main(page_url, target_tag, target_attr, os.path.join(folder_path, f'{file_name}.docx'))
                file_size = os.stat(os.path.join(folder_path, f'{file_name}.docx')).st_size // 1024
                if file_size < 35:
                    os.remove(os.path.join(folder_path, f'{file_name}.docx'))
                    logger.info(f'{Fore.RED}{file_name}.docx {file_size}KB{Style.RESET_ALL}')
                else:
                    logger.info(f'{Fore.BLUE}{file_name}.docx {file_size}KB{Style.RESET_ALL}')
                    data_map = {'标题': file_name, '目录': menu, '发布时间': fb_time, '来源': ly_name, '地址': page_url}
                    data = pd.concat([data, pd.DataFrame([data_map])], ignore_index=True)
            except Exception as e:
                logger.error(f"Error processing page {page_url}: {e}")
    except Exception as e:
        logger.error(f"Error fetching URL {url}: {e}")

write_to_excel(file_path=excel_file_path, dataframe=data, new_sheet=notification_path)
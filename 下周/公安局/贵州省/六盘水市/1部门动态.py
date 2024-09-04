import random
import re
import os
import time
import lxml.html
import pandas as pd
import requests
from colorama import Fore, Style
from openpyxl import load_workbook
import logging

logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
}

def write_to_excel(file_path, dataframe, new_sheet):
    if not os.path.exists(file_path):
        dataframe.to_excel(file_path, sheet_name=new_sheet, index=False)
    else:
        book = load_workbook(file_path)
        if new_sheet in book.sheetnames:
            existing_data = pd.read_excel(file_path, sheet_name=new_sheet)
            updated_data = existing_data._append(dataframe)
        else:
            updated_data = dataframe
        with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            updated_data.to_excel(writer, sheet_name=new_sheet, index=False)

url_name = '六盘水市公安局'
notification_path = '部门动态'
base_path = f'E:\WorkingWord\马缕_公文爬取(7.8-7.12)\公安局\贵州省\\{url_name}'
excel_file_name = f'{url_name}.xlsx'
folder_path = os.path.join(base_path, notification_path)

if not os.path.exists(folder_path):
    os.makedirs(folder_path)
    logger.info(f'Created directory: {folder_path}')

excel_file_path = os.path.join(base_path, excel_file_name)
data = pd.DataFrame()

page_list_xpath = '//div[@class="listmenubottom"]/ul/li/a/@href'
tit_xpath = '//*[@id="c"]/div[1]/text()'
menu_xpath = '//div[@class="path"]//a//text()'
fb_time_xpath = '//*[@id="c"]/div[3]/span[1]/text()'
ly_name_xpath = '//*[@id="c"]/div[3]/span[2]/text()'

for i in range(1, 63):
    url = f'https://gaj.gzlps.gov.cn/gzdt/bmdt/index_{i}.html'
    logger.info(f'URL: {url}')
    response = requests.get(url=url, headers=headers)
    response.raise_for_status()
    response.encoding = 'utf-8'
    tree = lxml.html.fromstring(response.text)
    page_list = tree.xpath(page_list_xpath)

    for page in page_list:
        try:
            time.sleep(random.uniform(1, 2))
            page_url = page
            page_response = requests.get(url=page_url, headers=headers)
            page_response.encoding = 'utf-8'
            page_tree = lxml.html.fromstring(page_response.text)

            tit = re.sub(r'[\r\n\s\t]*', '', page_tree.xpath(tit_xpath)[0])
            menu = '>'.join(page_tree.xpath(menu_xpath))
            fb_time = page_tree.xpath(fb_time_xpath)
            ly_name = page_tree.xpath(ly_name_xpath)
            file_path = f'{folder_path}\\{tit}.docx'

            with open(file_path, 'a+', encoding='utf-8') as file:
                file.write(tit + '\n')
                for p_text in page_tree.xpath('//font[@id="Zoom"]//p'):
                    content = ''.join(p_text.xpath('.//text()'))
                    file.write(content + '\n')

            file_size = os.stat(file_path).st_size // 1024
            if file_size == 0:
                os.remove(file_path)
                logger.info(f'{Fore.RED}{tit}.docx {file_size}KB{Style.RESET_ALL}')
            else:
                logger.info(f'{Fore.BLUE}{tit}.docx {file_size}KB{Style.RESET_ALL}')
                dataMap = {'标题': tit, '目录': menu, '发布时间': fb_time, '来源': ly_name, '地址': page_url}
                data = data._append(dataMap, ignore_index=True)

        except Exception as e:
            logger.error(f'Error processing page {page_url}: {e}')
            continue

write_to_excel(file_path=excel_file_path, dataframe=data, new_sheet=notification_path)

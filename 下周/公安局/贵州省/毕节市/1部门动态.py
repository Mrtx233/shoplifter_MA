import random
import re
import os
import time
import lxml.html
import pandas as pd
import requests
from colorama import Fore, Style
from bs4 import BeautifulSoup
from docx import Document
import logging
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
}

def write_to_excel(file_path, dataframe, new_sheet):
    if not os.path.exists(file_path):
        dataframe.to_excel(file_path, sheet_name=new_sheet, index=False)
    else:
        book = load_workbook(file_path)
        if new_sheet in book.sheetnames:
            existing_data = pd.read_excel(file_path, sheet_name=new_sheet)
            updated_data = pd.concat([existing_data, dataframe], ignore_index=True)
        else:
            updated_data = dataframe
        with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            updated_data.to_excel(writer, sheet_name=new_sheet, index=False)

url_name = '毕节市公安局'
notification_path = '部门动态'
base_path = f'E:\WorkingWord\马缕_公文爬取(7.8-7.12)\公安局\贵州省\\{url_name}'
excel_file_name = f'{url_name}.xlsx'
folder_path = os.path.join(base_path, notification_path)
excel_file_path = os.path.join(base_path, excel_file_name)

if not os.path.exists(folder_path):
    os.makedirs(folder_path)
    logger.info(f'Created directory: {folder_path}')

data = pd.DataFrame()

base_url = 'https://www.bijie.gov.cn/bm/bjsgaj/dt/bmdt/'
page_list_xpath = '//div[@class="PageMainBox aBox"]/ul/li/a/@href'
tit_xpath = '/html/body/div[9]/div[2]/div[2]/div[1]/text()'
menu_xpath = '//div[@class="Address Box"]//a//text()'
fb_time_xpath = '/html/body/div[9]/div[2]/div[2]/div[2]/div[1]/p[1]/span[2]/text()'
ly_name_xpath = '/html/body/div[9]/div[2]/div[2]/div[2]/div[1]/p[1]/span[1]/text()'

# 循环抓取页面数据
for i in range(1, 55):
    if i == 0:
        url = base_url
    else:
        url = f'{base_url}index_{i}.html'

    logger.info(f'URL: {url}')
    try:
        response = requests.get(url=url, headers=headers)
        response.raise_for_status()
        response.encoding = 'utf-8'
        tree = lxml.html.fromstring(response.text)
        page_list = tree.xpath(page_list_xpath)

        for page in page_list:
            try:
                time.sleep(random.uniform(1, 1.5))
                page_url = page
                page_response = requests.get(url=page_url, headers=headers)
                page_response.raise_for_status()
                page_response.encoding = 'utf-8'
                page_tree = lxml.html.fromstring(page_response.text)

                file_name = re.sub(r'[\r\n\s\t]*', '', page_tree.xpath(tit_xpath)[0])
                menu = '>'.join(page_tree.xpath(menu_xpath))
                fb_time = page_tree.xpath(fb_time_xpath)
                ly_name = page_tree.xpath(ly_name_xpath)
                # print(file_name,menu,fb_time,ly_name)

                html_content = page_response.content
                soup = BeautifulSoup(html_content, 'html.parser')
                target_div = soup.find('div', id='Zoom')
                text_content = target_div.get_text()

                doc = Document()
                doc.add_paragraph(text_content)
                save_path = os.path.join(folder_path, f'{file_name}.docx')
                doc.save(save_path)

                file_size = os.stat(save_path).st_size // 1024
                if file_size == 0:
                    os.remove(save_path)
                    logger.info(f'{Fore.RED}{file_name}.docx {file_size}KB{Style.RESET_ALL}')
                else:
                    logger.info(f'{Fore.BLUE}{file_name}.docx {file_size}KB{Style.RESET_ALL}')
                    data_map = {'标题': file_name, '目录': menu, '发布时间': fb_time, '来源': ly_name, '地址': page_url}
                    data = pd.concat([data, pd.DataFrame([data_map])], ignore_index=True)

            except Exception as e:
                logger.error(f"Error processing page {page_url}: {e}")

    except Exception as e:
        logger.error(f"Error fetching URL {url}: {e}")

write_to_excel(file_path=excel_file_path, dataframe=data, new_sheet=notification_path)

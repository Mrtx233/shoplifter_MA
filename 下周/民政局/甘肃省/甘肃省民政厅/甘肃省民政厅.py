import random
import re
import os
import time
import lxml.html
import pandas as pd
import requests
from colorama import Fore, Style
from openpyxl import load_workbook
import logging

logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
}

def write_to_excel(file_path, dataframe, new_sheet):
    if not os.path.exists(file_path):
        dataframe.to_excel(file_path, sheet_name=new_sheet, index=False)
    else:
        book = load_workbook(file_path)
        if new_sheet in book.sheetnames:
            existing_data = pd.read_excel(file_path, sheet_name=new_sheet)
            updated_data = pd.concat([existing_data, dataframe], ignore_index=True)
        else:
            updated_data = dataframe
        with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            updated_data.to_excel(writer, sheet_name=new_sheet, index=False)

base_url = 'https://sfj.gz.gov.cn/xxgk/xxgkml/gzdt/sfyw/index.html'
url_name = '甘肃省民政厅'
notification_path = '甘肃要闻'
base_path = f'E:\WorkingWord\马缕_公文爬取(7.8-7.12)\民政局\甘肃省\\{url_name}'
excel_file_name = f'{url_name}.xlsx'

page_list_xpath = '/html/body/div[4]/div/div[1]/div[2]/div[2]/div[1]/a/@href'
# tit_xpath = '//*[@id="ConBox"]/div[2]/h1/text()'
# menu_xpath = '//div[@class="dh mb5 b-ltrb"]//a//text()'
# fb_time_xpath = '//*[@id="ConBox"]/div[2]/div/div[1]/span[2]/text()'
# ly_name_xpath = '//*[@id="ConBox"]/div[2]/div/div[1]/span[1]/text()'



for i in range(0, 1):
    if i == 0:
        url = base_url
    else:
        url = f'{base_url}index_{i}.html'

    logger.info(f'URL: {url}')
    response = requests.get(url=url, headers=headers)
    response.raise_for_status()
    response.encoding = 'utf-8'
    tree = lxml.html.fromstring(response.text)
    page_list = tree.xpath(page_list_xpath)

    for page in page_list:
        try:
            page_url = page
            page_response = requests.get(url=page_url, headers=headers)
            page_response.encoding = 'utf-8'
            page_tree = lxml.html.fromstring(page_response.text)
            print(page)
        except Exception as e:
            logger.error(f'Error processing page {page_url}: {e}')
            continue

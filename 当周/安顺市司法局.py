import random  # 导入随机模块，用于生成随机数
import re  # 导入正则表达式模块，用于处理字符串
import os  # 导入操作系统模块，用于文件操作
import time  # 导入时间模块，用于暂停和时间相关操作
import html  # 导入HTML模块，用于处理HTML实体
import lxml.html  # 导入lxml模块中的HTML子模块，用于解析HTML
import pandas as pd  # 导入pandas库并使用别名pd，用于数据处理
import requests  # 导入requests库，用于发送HTTP请求
from colorama import Fore, Style  # 导入colorama库中的Fore和Style，用于彩色输出
from bs4 import BeautifulSoup  # 导入BeautifulSoup库，用于解析HTML
from docx import Document  # 导入docx库中的Document类，用于创建Word文档
import logging  # 导入日志模块，用于记录日志
from openpyxl import load_workbook  # 导入openpyxl中的load_workbook，用于处理Excel文件

# 配置日志格式和级别
logging.basicConfig(level=logging.INFO, format='%(asctime)s >>> %(message)s')
logger = logging.getLogger(__name__)  # 获取logger实例

# 设置请求头，使用固定的User-Agent
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
}

# 定义函数将DataFrame写入Excel
def write_to_excel(file_path, dataframe, new_sheet):
    # 检查文件是否存在
    if not os.path.exists(file_path):
        # 如果文件不存在，直接创建新的Excel文件
        dataframe.to_excel(file_path, sheet_name=new_sheet, index=False)
    else:
        # 如果文件存在，加载工作簿
        book = load_workbook(file_path)
        # 检查工作表是否存在
        if new_sheet in book.sheetnames:
            # 如果工作表存在，加载到pandas DataFrame中
            existing_data = pd.read_excel(file_path, sheet_name=new_sheet)
            # 追加新的数据
            updated_data = existing_data._append(dataframe)
        else:
            # 如果工作表不存在，直接使用新的数据
            updated_data = dataframe
        # 将数据写入Excel文件
        with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            updated_data.to_excel(writer, sheet_name=new_sheet, index=False)

# 基础路径
base_path = r'E:\WorkingWord\马缕_公文爬取(7.3-7.5)\安顺市司法局'
# 通知公告路径
notification_path = '依法治市'
# Excel文件名
excel_file_name = '安顺市司法局.xlsx'
# 合并基础路径和通知公告路径
folder_path = os.path.join(base_path, notification_path)
# 构建Excel文件路径
excel_file_path = os.path.join(base_path, excel_file_name)

data = pd.DataFrame()  # 初始化空的DataFrame来存储数据
for i in range(2, 3):  # 循环遍历第11到20页
    url = f'https://sfj.anshun.gov.cn/2021gb/ywgz_5890798/yfzs1/'  # 构建URL
    logger.info(f'URL: {url}')  # 记录当前URL

    response = requests.get(url=url, headers=headers)  # 发起HTTP GET请求
    response.raise_for_status()  # 如果请求失败则抛出异常
    response.encoding = 'utf-8'  # 设置响应的编码为utf-8
    tree = lxml.html.fromstring(response.text)  # 解析HTML
    page_list = tree.xpath('//ul[@class="main_lis"]/li/a/@href')  # 提取每个页面的链接

    for page in page_list:  # 循环遍历每个页面链接
        try:
            # time.sleep(random.uniform(1, 1.5))  # 随机等待1到1.5秒，避免被封IP
            # page_url = 'https://gaj.weinan.gov.cn/' + re.sub(r'\./', '/', page)  # 构建页面URL
            page_url=page
            page_response = requests.get(url=page_url, headers=headers)  # 发起HTTP GET请求获取页面内容
            page_response.encoding = 'utf-8'  # 设置响应的编码为utf-8
            page_tree = lxml.html.fromstring(page_response.text)  # 解析HTML

            tit = page_tree.xpath('/html/body/div[13]/div[2]/div/div[1]/h1/text()')[0]  # 提取标题
            tit = re.sub(r'[\r\n\s\t]*', '', tit)  # 清理标题中的空白字符
            menu = '>'.join(page_tree.xpath('//div[@class="current"]//a//text()'))  # 提取目录
            fb_time = page_tree.xpath('/html/body/div[11]/div[2]/div/div[1]/p/span[1]/text()')  # 提取发布时间
            ly_name = page_tree.xpath('/html/body/div[11]/div[2]/div/div[1]/p/span[2]/text()')  # 提取来源

            file_path = f'{folder_path}\\{tit}.docx'  # 构建文件路径
            file = open(file_path, 'a+', encoding='utf-8')  # 以追加模式打开文件
            file.write(tit + '\n')  # 写入标题

            content_p = page_tree.xpath('//div[@class="detailsMain"]/div/p')  # 提取内容段落
            if len(content_p) > 0:  # 如果内容段落存在
                for p_text in content_p:  # 循环遍历每个段落
                    content = ''.join(p_text.xpath('./span/text()'))  # 提取段落中的文本
                    file.write(content + '\n')  # 写入文件

            file.close()  # 关闭文件
            file_size = os.stat(file_path).st_size % 1024  # 获取文件大小（单位：KB）
            if file_size == 0:  # 如果文件大小为0KB，删除文件
                os.remove(file_path)
                logger.info(f'{Fore.RED}{tit}.docx {file_size}KB{Style.RESET_ALL}')
            else:  # 否则记录文件大小并追加数据到DataFrame
                logger.info(f'{Fore.BLUE}{tit}.docx {file_size}KB{Style.RESET_ALL}')
                dataMap = {'标题': tit, '目录': menu, '发布时间': fb_time, '来源': ly_name, '地址': page_url}
                data = data._append(dataMap, ignore_index=True)

        except Exception as e:  # 捕获所有异常
            logger.error(f'Error processing page {page_url}: {e}')  # 记录错误信息
            continue

# 将DataFrame写入Excel文件
write_to_excel(file_path=excel_file_path, dataframe=data, new_sheet=notification_path)

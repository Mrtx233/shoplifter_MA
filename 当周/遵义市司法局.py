import random  # 导入随机模块，用于生成随机数
import re  # 导入正则表达式模块，用于处理字符串
import os  # 导入操作系统模块，用于文件操作
import time  # 导入时间模块，用于暂停和时间相关操作
import html  # 导入HTML模块，用于处理HTML实体
import lxml.html  # 导入lxml模块中的HTML子模块，用于解析HTML
import pandas as pd  # 导入pandas库并使用别名pd，用于数据处理
import requests  # 导入requests库，用于发送HTTP请求
from colorama import Fore, Style  # 导入colorama库中的Fore和Style，用于彩色输出
from bs4 import BeautifulSoup  # 导入BeautifulSoup库，用于解析HTML
from docx import Document  # 导入docx库中的Document类，用于创建Word文档
import logging  # 导入日志模块，用于记录日志
from openpyxl import load_workbook  # 导入openpyxl中的load_workbook，用于处理Excel文件

# 日志配置，设定日志级别为INFO，日志格式为简单的消息内容
logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)  # 获取logger对象

# 设置请求头，模拟浏览器访问
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
}

# 定义将数据写入Excel文件的函数
def write_to_excel(file_path, dataframe, new_sheet):
    # 检查文件是否存在
    if not os.path.exists(file_path):
        # 如果文件不存在，直接创建新的Excel文件
        dataframe.to_excel(file_path, sheet_name=new_sheet, index=False)
    else:
        # 如果文件存在，加载工作簿
        book = load_workbook(file_path)
        # 检查工作表是否存在
        if new_sheet in book.sheetnames:
            # 如果工作表存在，加载到pandas DataFrame中
            existing_data = pd.read_excel(file_path, sheet_name=new_sheet)
            # 追加新的数据
            updated_data = existing_data._append(dataframe)
        else:
            # 如果工作表不存在，直接使用新的数据
            updated_data = dataframe
        # 将数据写入Excel文件
        with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            updated_data.to_excel(writer, sheet_name=new_sheet, index=False)

# 基础路径
base_path = r'E:\WorkingWord\马缕_公文爬取(7.3-7.5)\遵义市司法局'
# 通知公告路径
notification_path = '区县动态'
# Excel文件名
excel_file_name = '遵义市司法局.xlsx'

# 合并基础路径和通知公告路径
folder_path = os.path.join(base_path, notification_path)
# 构建Excel文件路径
excel_file_path = os.path.join(base_path, excel_file_name)

data = pd.DataFrame()  # 初始化一个空的DataFrame用于存储数据

for i in range(1, 6):  # 循环生成分页URL
    url = f'https://sfj.zunyi.gov.cn/xwdt/tzgg/index_{i}.html'  # 构建URL
    logger.info(f'URL: {url}')  # 记录当前URL

    try:
        # 发送GET请求获取网页内容
        response = requests.get(url=url, headers=headers)
        response.raise_for_status()  # 如果请求失败则抛出异常
        response.encoding = 'utf-8'  # 设置响应的编码为utf-8
        tree = lxml.html.fromstring(response.text)  # 解析HTML
        page_list = tree.xpath('//ul[@class="NewsList"]/li/a/@href')  # 提取每个页面的链接

        for page in page_list:  # 遍历每个页面链接
            time.sleep(random.uniform(1, 1.5))  # 随机等待1到1.5秒，避免被封IP
            page_url = page  # 获取页面URL
            # page_url = 'https://sfj.guiyang.gov.cn/xwzx/zhyw_5616653' + re.sub(r'\./', '/', page)
            try:
                # 发送GET请求获取页面内容
                page_response = requests.get(url=page_url, headers=headers)
                page_response.raise_for_status()  # 如果请求失败则抛出异常
                page_response.encoding = 'utf-8'  # 设置响应的编码为utf-8
                page_tree = lxml.html.fromstring(page_response.text)  # 解析HTML

                # 提取文件名，去除多余的空格和换行
                file_name = page_tree.xpath('/html/body/div[9]/div[3]/div[1]/div[1]/text()')[0]
                file_name = re.sub(r'[\r\n\s\t]*', '', file_name)
                # 提取目录
                menu = '>'.join(page_tree.xpath('//div[@class="w1200 Box auto"]//a//text()'))
                # 提取发布时间
                fb_time = page_tree.xpath('/html/body/div[9]/div[3]/div[1]/div[2]/div[1]/p[1]/span[2]/text()')
                # 提取来源
                ly_name = page_tree.xpath('/html/body/div[9]/div[3]/div[1]/div[2]/div[1]/p[1]/span[1]/text()')

                # 获取页面内容
                html_content = page_response.content
                soup = BeautifulSoup(html_content, 'html.parser')  # 解析HTML内容

                # 查找目标div
                target_div = soup.find('div', id='Zoom')
                text_content = target_div.get_text()  # 获取div的文本内容

                # 创建Word文档对象
                doc = Document()
                doc.add_paragraph(text_content)  # 添加文本内容到文档

                save_path = os.path.join(folder_path, f'{file_name}.docx')  # 构建保存路径

                # 保存Word文档
                doc.save(save_path)

                # 获取文件大小（以KB为单位）
                file_size = os.stat(save_path).st_size // 1024
                if file_size == 0:  # 如果文件大小为0，删除文件
                    os.remove(save_path)
                    logger.info(f'{Fore.RED}{file_name}.docx {file_size}KB{Style.RESET_ALL}')  # 记录文件信息
                else:  # 如果文件大小不为0，记录文件信息并保存数据
                    logger.info(f'{Fore.BLUE}{file_name}.docx {file_size}KB{Style.RESET_ALL}')
                    data_map = {'标题': file_name, '目录': menu, '发布时间': fb_time, '来源': ly_name, '地址': page_url}
                    data = data._append(data_map, ignore_index=True)

            except Exception as e:  # 捕获页面处理过程中的异常
                logger.error(f"Error processing page {page_url}: {e}")

    except Exception as e:  # 捕获获取页面列表过程中的异常
        logger.error(f"Error fetching URL {url}: {e}")

# 将数据写入Excel文件
write_to_excel(file_path=excel_file_path, dataframe=data, new_sheet=notification_path)

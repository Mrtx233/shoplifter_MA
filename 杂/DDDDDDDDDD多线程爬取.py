import random
import re
import os
import time
import lxml.html
import pandas as pd
import requests
from colorama import Fore, Style
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
import logging

logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
}


def write_to_excel(file_path, dataframe, new_sheet):
    if not os.path.exists(file_path):
        dataframe.to_excel(file_path, sheet_name=new_sheet, index=False)
    else:
        book = load_workbook(file_path)
        if new_sheet in book.sheetnames:
            existing_data = pd.read_excel(file_path, sheet_name=new_sheet)
            updated_data = pd.concat([existing_data, dataframe], ignore_index=True)
        else:
            updated_data = dataframe
        with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            updated_data.to_excel(writer, sheet_name=new_sheet, index=False)



def fetch_page_data(base_url, notification_path):
    A_url = 'https://jiankang.gmw.cn/'
    base_path = f'E:\\WorkingWord\\马缕_新闻稿采集(7.17-7.19)\\中央新闻网站\\光明网'
    excel_file_name = '光明网.xlsx'
    folder_path = os.path.join(base_path, notification_path)
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
        logger.info(f'Created directory: {folder_path}')
    excel_file_path = os.path.join(base_path, excel_file_name)
    data = pd.DataFrame()

    page_list_xpath = '/html/body/div[6]/div[1]/div[2]/ul/li/a/@href'
    tit_xpath = '/html/body/div[6]/div[1]/h1/text()'
    alt_tit_xpath = '//title/text()'  # 备用XPath
    menu_xpath = '//div[@class="g-crumbs"]//a//text()'
    fb_time_xpath = '/html/body/div[6]/div[1]/div/div[1]/span[2]/text()'
    ly_name_xpath = '/html/body/div[6]/div[1]/div/div[1]/span[1]/a/text()'

    for i in range(1, 11):
        try:
            if i == 1:
                url = f'{base_url}.htm'
            else:
                url = f'{base_url}_{i}.htm'
            logger.info(f'URL: {url}')
            response = requests.get(url=url, headers=headers)
            response.raise_for_status()
            response.encoding = 'utf-8'
            tree = lxml.html.fromstring(response.text)
            page_list = tree.xpath(page_list_xpath)

            for page in page_list:
                try:
                    time.sleep(random.uniform(0.5, 0.8))
                    if re.match(r'^https', page):
                        page_url = page
                    else:
                        page_url = A_url + re.sub(r'\./', '/', page)
                    page_response = requests.get(url=page_url, headers=headers)
                    page_response.encoding = 'utf-8'
                    page_tree = lxml.html.fromstring(page_response.text)

                    # 提取标题，如果初始tit_xpath为空则使用备用alt_tit_xpath
                    tit = page_tree.xpath(tit_xpath)
                    if not tit:
                        tit = page_tree.xpath(alt_tit_xpath)
                    tit = re.sub(r'[\r\n\s\t\|]*', '', tit[0]) if tit else '无标题'

                    menu = '>'.join(page_tree.xpath(menu_xpath))
                    fb_time = page_tree.xpath(fb_time_xpath)
                    ly_name = page_tree.xpath(ly_name_xpath)
                    file_path = f'{folder_path}\\{tit}.docx'

                    with open(file_path, 'a+', encoding='utf-8') as file:
                        file.write(tit + '\n')
                        for p_text in page_tree.xpath('//*[@id="article_inbox"]/div[5]//p'):
                            if p_text.xpath('.//text()') != '':
                                content = ''.join(p_text.xpath('.//text()'))
                                file.write(content + '\n')
                            else:
                                content = ''.join(p_text.xpath('./span/text()'))
                                file.write(content + '\n')

                    file_size = os.stat(file_path).st_size // 1024
                    if file_size == 0:
                        os.remove(file_path)
                        logger.info(f'{Fore.RED}{tit}.docx {file_size}KB{Style.RESET_ALL}')
                    else:
                        logger.info(f'{Fore.BLUE}{tit}.docx {file_size}KB{Style.RESET_ALL}')
                        dataMap = {'标题': tit, '目录': menu, '发布时间': fb_time, '来源': ly_name, '地址': page_url}
                        data = pd.concat([data, pd.DataFrame([dataMap])], ignore_index=True)

                except Exception as e:
                    logger.error(f'Error processing page {page_url}: {e}')
                    continue
        except Exception as e:
            logger.error(f'Error processing page {url}: {e}')
            continue

    write_to_excel(file_path=excel_file_path, dataframe=data, new_sheet=notification_path)
    time.sleep(10)


if __name__ == '__main__':
    base_urls = [
        'https://jiankang.gmw.cn/node_12215',
        'https://jiankang.gmw.cn/node_12206',
        'https://jiankang.gmw.cn/node_12212',
        'https://jiankang.gmw.cn/node_12207',
        'https://jiankang.gmw.cn/node_12202',
        'https://jiankang.gmw.cn/node_12211',
        'https://jiankang.gmw.cn/node_12295'
    ]

    notification_paths = [
        '健康频道-要闻',
        '健康频道-健康常识',
        '健康频道-美容美体',
        '健康频道-营养保健',
        '健康频道-养生名人堂',
        '健康频道-品牌活动',
        '健康频道-健康专区'
    ]

    with ThreadPoolExecutor(max_workers=6) as executor:
        futures = []
        for base_url, notification_path in zip(base_urls, notification_paths):
            futures.append(executor.submit(fetch_page_data, base_url, notification_path))

        for future in as_completed(futures):
            try:
                future.result()
            except Exception as e:
                logger.error(f'Error in thread: {e}')

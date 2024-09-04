import time
import random
import re
import os
import requests
import lxml.html
import pandas as pd
from colorama import Fore, Style
from openpyxl import load_workbook
import logging
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
}

def write_to_excel(file_path, dataframe, new_sheet):
    if not os.path.exists(file_path):
        dataframe.to_excel(file_path, sheet_name=new_sheet, index=False)
    else:
        book = load_workbook(file_path)
        if new_sheet in book.sheetnames:
            existing_data = pd.read_excel(file_path, sheet_name=new_sheet)
            updated_data = pd.concat([existing_data, dataframe], ignore_index=True)
        else:
            updated_data = dataframe
        with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            updated_data.to_excel(writer, sheet_name=new_sheet, index=False)

def fetch_links_with_more_button(url, target_xpath, button_class, pause_time=5, max_clicks=10):
    """
    打开指定的URL，点击“查看更多”按钮以加载更多内容，获取所有指定XPath的<a>标签链接，并打印所有链接。

    参数:
    - url: 要打开的网页的URL
    - target_xpath: 用于定位<a>标签的XPath
    - button_class: “查看更多”按钮的class
    - pause_time: 点击按钮后等待的时间
    - max_clicks: 最大点击次数

    返回:
    - target_xpath: 传入的XPath
    """
    # 设置Selenium WebDriver
    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
    driver.get(url)

    def click_more_button(driver, button_class, pause_time, max_clicks):
        """
        点击“查看更多”按钮以加载更多内容
        """
        clicks = 0
        while clicks < max_clicks:
            try:
                # 查找“查看更多”按钮并点击
                more_button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.CLASS_NAME, button_class))
                )
                more_button.click()
                time.sleep(pause_time)  # 等待内容加载完成
                clicks += 1
                print(f"点击“查看更多”按钮，当前点击次数: {clicks}")
            except Exception as e:
                print(f"点击按钮时发生错误: {e}")
                break

    def get_all_links(driver, xpath):
        """
        提取页面上所有符合XPath的<a>标签链接
        """
        elements = WebDriverWait(driver, 20).until(
            EC.presence_of_all_elements_located((By.XPATH, xpath))
        )
        links = set()
        for element in elements:
            href = element.get_attribute('href')
            if href:
                links.add(href)
        return links

    # 调用点击按钮的函数
    click_more_button(driver, button_class, pause_time, max_clicks)

    # 提取页面上所有的<a>标签链接
    final_links = get_all_links(driver, target_xpath)

    # 打印所有的<a>标签链接数量
    print(f"\n总共收集到的<a>标签数量: {len(final_links)}")
    print("所有<a>标签的链接:")
    for link in final_links:
        print(link)

    # 关闭浏览器
    driver.quit()

    # 返回target_xpath
    return final_links


# 使用示例
url = 'http://www.jjckb.cn/yw.htm'
target_xpath = '/html/body/div[1]/div[3]/ul/li/a'
button_class = 'xpage-more-btn'  # “查看更多”按钮的class
returned_final_links = fetch_links_with_more_button(url, target_xpath, button_class)

url_name = '经济参考网'
notification_path = '要闻NEWS'
base_path = f'E:\\WorkingWord\\马缕_公文爬取(7.15-7.19)\\{url_name}'
excel_file_name = f'{url_name}.xlsx'

tit_xpath = '/html/body/div[2]/div[2]/div[1]/div[1]/text()'
menu_xpath = '//div[@class="breadcrumb"]//a//text()'
fb_time_xpath = '/html/body/div[2]/div[2]/div[1]/div[2]/span[2]/text()'
ly_name_xpath = '/html/body/div[2]/div[2]/div[1]/div[2]/span[1]/text()'

folder_path = os.path.join(base_path, notification_path)
if not os.path.exists(folder_path):
    os.makedirs(folder_path)
    logger.info(f'Created directory: {folder_path}')
excel_file_path = os.path.join(base_path, excel_file_name)

data = pd.DataFrame()

for page in returned_final_links:
    try:
        time.sleep(random.uniform(1, 1.5))
        page_url = page
        page_response = requests.get(url=page_url, headers=headers)
        page_response.encoding = 'utf-8'
        page_tree = lxml.html.fromstring(page_response.text)

        tit = re.sub(r'[\r\n\s\t]*', '', page_tree.xpath(tit_xpath)[0])
        menu = '>'.join(page_tree.xpath(menu_xpath))
        fb_time = page_tree.xpath(fb_time_xpath)
        ly_name = page_tree.xpath(ly_name_xpath)
        file_path = f'{folder_path}\\{tit}.docx'

        with open(file_path, 'a+', encoding='utf-8') as file:
            file.write(tit + '\n')
            for p_text in page_tree.xpath('/html/body/div[2]/div[2]/div[1]/div[3]/div[2]/p'):
                if p_text.xpath('.//text()') != '':
                    content = ''.join(p_text.xpath('.//text()'))
                    file.write(content + '\n')
                else:
                    content = ''.join(p_text.xpath('./span/text()'))
                    file.write(content + '\n')

        file_size = os.stat(file_path).st_size // 1024
        if file_size == 0:
            os.remove(file_path)
            logger.info(f'{Fore.RED}{tit}.docx {file_size}KB{Style.RESET_ALL}')
        else:
            logger.info(f'{Fore.BLUE}{tit}.docx {file_size}KB{Style.RESET_ALL}')
            dataMap = {'标题': tit, '目录': menu, '发布时间': fb_time, '来源': ly_name, '地址': page_url}
            data = pd.concat([data, pd.DataFrame([dataMap])], ignore_index=True)

    except Exception as e:
        logger.error(f'Error processing page {page_url}: {e}')
        continue

write_to_excel(file_path=excel_file_path, dataframe=data, new_sheet=notification_path)

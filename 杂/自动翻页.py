import time
import random
import re
import os
import requests
import lxml.html
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from colorama import Fore, Style
from openpyxl import load_workbook
import logging

logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
}

def write_to_excel(file_path, dataframe, new_sheet):
    if not os.path.exists(file_path):
        dataframe.to_excel(file_path, sheet_name=new_sheet, index=False)
    else:
        book = load_workbook(file_path)
        if new_sheet in book.sheetnames:
            existing_data = pd.read_excel(file_path, sheet_name=new_sheet)
            updated_data = pd.concat([existing_data, dataframe], ignore_index=True)
        else:
            updated_data = dataframe
        with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            updated_data.to_excel(writer, sheet_name=new_sheet, index=False)
def extract_links_from_page(driver, link_xpath):
    elements = driver.find_elements(By.XPATH, link_xpath)
    links = [element.get_attribute('href') for element in elements]
    return links

def scrape_links(url, link_xpath, next_button_xpath, click_times):
    driver = webdriver.Chrome()
    all_links = []

    try:
        driver.get(url)

        for _ in range(click_times):
            time.sleep(2)  # 等待页面加载

            # 提取当前页面的链接
            links = extract_links_from_page(driver, link_xpath)
            all_links.extend(links)

            # 点击下一页按钮
            next_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, next_button_xpath))
            )
            next_button.click()

    finally:
        driver.quit()

    return all_links

url = 'https://www.cenews.com.cn/column.html?cid=165'
link_xpath = '//*[@id="app"]/div[3]/div/div[1]/div/div[2]/ul/li/div[1]/div[1]/h2/a'
next_button_xpath = '//*[@id="app"]/div/div[2]/div[3]/div[2]/div/div/button[2]/i'
click_times = 0

all_links = scrape_links(url, link_xpath, next_button_xpath, click_times)

url_name = '中国商报网'
notification_path = '商业'
base_path = f'E:\\WorkingWord\\马缕_新闻稿采集(7.22-7.26)\\中央新闻网站\\{url_name}'
excel_file_name = f'{url_name}.xlsx'

tit_xpath = '//*[@id="app"]/div[2]/div/div[1]/div/div/div[2]/div[1]/h5/text()'
menu_xpath = '//div[@class="locationBox el-row"]//a//text()'
fb_time_xpath = '//*[@id="app"]/div[2]/div/div[1]/div/div/div[2]/div[1]/p/span[1]/text()'
ly_name_xpath = '//*[@id="app"]/div[2]/div/div[1]/div/div/div[2]/div[1]/p/span[2]/text()'

folder_path = os.path.join(base_path, notification_path)
if not os.path.exists(folder_path):
    os.makedirs(folder_path)
    logger.info(f'Created directory: {folder_path}')
excel_file_path = os.path.join(base_path, excel_file_name)

data = pd.DataFrame()

for page in all_links:
    try:
        time.sleep(random.uniform(1, 1.5))
        page_url = page
        page_response = requests.get(url=page_url, headers=headers)
        page_response.encoding = 'utf-8'
        page_tree = lxml.html.fromstring(page_response.text)

        tit = re.sub(r'[\r\n\s\t]*', '', page_tree.xpath(tit_xpath)[0])
        menu = '>'
        fb_time = page_tree.xpath(fb_time_xpath)
        ly_name = page_tree.xpath(ly_name_xpath)
        file_path = f'{folder_path}\\{tit}.docx'

        with open(file_path, 'a+', encoding='utf-8') as file:
            file.write(tit + '\n')
            for p_text in page_tree.xpath('//*[@id="app"]/div[2]/div/div[1]/div/div/div[2]/div[2]//p'):
                if p_text.xpath('.//text()') != '':
                    content = ''.join(p_text.xpath('.//text()'))
                    file.write(content + '\n')
                else:
                    content = ''.join(p_text.xpath('./span/text()'))
                    file.write(content + '\n')

        file_size = os.stat(file_path).st_size // 1024
        if file_size == 0:
            os.remove(file_path)
            logger.info(f'{Fore.RED}{tit}.docx {file_size}KB{Style.RESET_ALL}')
        else:
            logger.info(f'{Fore.BLUE}{tit}.docx {file_size}KB{Style.RESET_ALL}')
            dataMap = {'标题': tit, '目录': menu, '发布时间': fb_time, '来源': ly_name, '地址': page_url}
            data = pd.concat([data, pd.DataFrame([dataMap])], ignore_index=True)

    except Exception as e:
        logger.error(f'Error processing page {page_url}: {e}')
        continue

write_to_excel(file_path=excel_file_path, dataframe=data, new_sheet=notification_path)
